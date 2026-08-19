from datetime import UTC, datetime
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pytest

from agentgate.case.excel_import_export import (
    DatasetExcelValidationError,
    build_excel,
    parse_excel,
)
from agentgate.domain import (
    Case,
    CaseCategory,
    CaseDifficulty,
    CaseTurn,
    DatasetVersion,
    DatasetVersionStatus,
    Equals,
    MatchesPattern,
    OutputExpectation,
    StateExpectation,
    ToolArgumentExpectation,
)


HEADERS = (
    "case_id", "case_name", "case_description", "category", "difficulty",
    "tags_json", "initial_state_json", "turn_id", "turn_order", "input_json",
    "expected_skill", "expectations_json", "required_tools_json", "forbidden_tools_json",
    "policy_rules_json", "turn_notes",
)


def published_version_with_all_fields() -> DatasetVersion:
    case = Case(
        id="case-001",
        name="贷款审批",
        notes="多轮审批场景",
        category=CaseCategory.BOUNDARY,
        difficulty=CaseDifficulty.HARD,
        tags=("loan", "中文"),
        initial_state={"application": {"id": "A-1", "status": "new"}},
        turns=(
            CaseTurn(
                id="turn-001",
                input={"message": "申请贷款", "amount": 50000},
                expected_skill="loan_approval",
                expectations=(StateExpectation(
                    id="expect-state",
                    path="application.status",
                    condition=Equals(expected="pending"),
                    name="状态已更新",
                ),),
                required_tools=("credit_check",),
                forbidden_tools=("approve_loan",),
                policy_rules=("high_risk_requires_review",),
                notes="先收集资料",
            ),
            CaseTurn(
                id="turn-002",
                input={"application_id": "A-1", "risk": "high"},
                expected_skill="loan_approval",
                expectations=(
                    ToolArgumentExpectation(
                        id="expect-tool",
                        tool="request_human_review",
                        path="application_id",
                        occurrence="first",
                        condition=Equals(expected="A-1"),
                    ),
                    OutputExpectation(
                        id="expect-output",
                        path="message",
                        condition=MatchesPattern(pattern="审核"),
                    ),
                ),
                required_tools=("request_human_review",),
                notes="人工审核",
            ),
        ),
    )
    return DatasetVersion(
        id="version-001",
        dataset_id="dataset-001",
        dataset_name="贷款数据集",
        version=1,
        status=DatasetVersionStatus.PUBLISHED,
        cases=(case,),
        published_at=datetime(2026, 8, 19, tzinfo=UTC),
    )


def test_excel_round_trip_preserves_multiturn_case_and_ids():
    version = published_version_with_all_fields()

    content = build_excel(version)

    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Cases"]
    assert workbook.sheetnames == ["Cases"]
    assert tuple(next(sheet.iter_rows(values_only=True))) == HEADERS
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == ["case-001", "case-001"]
    assert [row[7] for row in rows] == ["turn-001", "turn-002"]
    assert [row[8] for row in rows] == [1, 2]
    assert parse_excel(content) == version.cases


def _workbook_bytes(rows=(), headers=HEADERS, sheet_name="Cases"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _valid_row(**changes):
    row = dict(zip(HEADERS, (
        "case-1", "Case", "", "positive", "medium", "[]", "{}", "turn-1", 1,
        '{"message":"hello"}', None, "[]", "[]", "[]", "[]", "",
    ), strict=True))
    row.update(changes)
    return tuple(row[header] for header in HEADERS)


def _issues(content):
    with pytest.raises(DatasetExcelValidationError) as error:
        parse_excel(content)
    return error.value.issues


def test_excel_reports_missing_sheet_and_header_locations():
    missing_sheet = _issues(_workbook_bytes(sheet_name="Other"))
    malformed_header = _issues(_workbook_bytes(headers=HEADERS[:-1]))

    assert missing_sheet[0].sheet == "Cases"
    assert missing_sheet[0].row is None
    assert missing_sheet[0].column is None
    assert malformed_header[0].sheet == "Cases"
    assert malformed_header[0].row == 1
    assert malformed_header[0].column == "turn_notes"


def test_excel_aggregates_blank_and_malformed_json_cells():
    issues = _issues(_workbook_bytes((
        _valid_row(input_json=None, tags_json="{"),
        _valid_row(case_id="case-2", turn_id="turn-2", expectations_json="["),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "input_json"), (2, "tags_json"), (3, "expectations_json"),
    }
    assert all(issue.sheet == "Cases" and issue.message for issue in issues)


def test_excel_reports_conflicting_case_fields_and_duplicate_turn_ids():
    issues = _issues(_workbook_bytes((
        _valid_row(),
        _valid_row(case_name="Changed", turn_order=2),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (3, "case_name"), (3, "turn_id"),
    }


def test_excel_reports_duplicate_and_noncontiguous_turn_orders():
    issues = _issues(_workbook_bytes((
        _valid_row(turn_id="turn-1", turn_order=1),
        _valid_row(turn_id="turn-2", turn_order=1),
        _valid_row(turn_id="turn-3", turn_order=3),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (3, "turn_order"), (4, "turn_order"),
    }


def test_excel_reports_invalid_case_enums_and_expectations():
    issues = _issues(_workbook_bytes((
        _valid_row(
            category="unknown",
            difficulty="impossible",
            expectations_json='[{"kind":"state"}]',
        ),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "category"), (2, "difficulty"), (2, "expectations_json"),
    }


def test_excel_reports_row_limit_and_invalid_xlsx_content():
    content = _workbook_bytes([
        _valid_row(turn_id=f"turn-{index}") for index in range(10_001)
    ])
    row_limit = _issues(content)
    invalid_xlsx = _issues(b"not an xlsx")

    assert row_limit[0].sheet == "Cases"
    assert row_limit[0].row == 10_002
    assert row_limit[0].column is None
    assert invalid_xlsx[0].sheet == "Cases"
    assert invalid_xlsx[0].row is None
    assert invalid_xlsx[0].column is None


def test_excel_reports_input_byte_limit_before_reading_workbook():
    issues = _issues(b"x" * (10 * 1024 * 1024 + 1))

    assert issues[0].sheet == "Cases"
    assert issues[0].row is None
    assert issues[0].column is None


def _corrupt_cases_xml(content):
    corrupted = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(corrupted, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            data = source.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = b"<worksheet"
            target.writestr(member, data)
    return corrupted.getvalue()


def test_excel_wraps_lazy_worksheet_xml_errors_as_import_issues():
    issues = _issues(_corrupt_cases_xml(_workbook_bytes()))

    assert issues[0].sheet == "Cases"
    assert issues[0].row is None
    assert issues[0].column is None
    assert issues[0].message.startswith("invalid XLSX content")


def test_excel_aggregates_model_errors_when_case_id_is_missing():
    issues = _issues(_workbook_bytes((
        _valid_row(
            case_id=None,
            category="unknown",
            difficulty="impossible",
            tags_json="{}",
            initial_state_json="[]",
            input_json="[]",
        ),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "case_id"),
        (2, "category"),
        (2, "difficulty"),
        (2, "tags_json"),
        (2, "initial_state_json"),
        (2, "input_json"),
    }
