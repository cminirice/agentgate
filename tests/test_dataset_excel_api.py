from io import BytesIO

import openpyxl
from fastapi.testclient import TestClient

from agentgate.case import build_excel
from agentgate.domain import Case, CaseTurn, DatasetVersion
from agentgate.server.application import create_app


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADERS = (
    "case_id", "case_name", "case_description", "category", "difficulty",
    "tags_json", "initial_state_json", "turn_id", "turn_order", "input_json",
    "expected_skill", "expectations_json", "required_tools_json", "forbidden_tools_json",
    "policy_rules_json", "turn_notes",
)


def _workbook_with_multiple_issues() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cases"
    sheet.append(HEADERS)
    sheet.append((
        "case-1", "Case", "", "positive", "medium", "{", "{}", "turn-1", 1,
        None, None, "[", "[]", "[]", "[]", "",
    ))
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _dataset_ids(client: TestClient) -> set[str]:
    return {item["id"] for item in client.get("/api/datasets").json()}


def test_excel_import_creates_draft_and_published_export_downloads(tmp_path):
    case = Case(
        id="imported-case",
        name="Imported Case",
        turns=(CaseTurn(id="imported-turn", input={"message": "hello"}),),
    )
    workbook = build_excel(DatasetVersion(dataset_id="source", cases=(case,)))

    with TestClient(create_app(tmp_path / "dataset-excel-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Imported Excel", "description": "From workbook"},
            files={
                "file": (
                    "import.xlsx",
                    workbook,
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert imported.status_code == 201
        payload = imported.json()
        dataset_id = payload["dataset"]["id"]
        assert dataset_id != "source"
        assert payload["dataset"]["name"] == "Imported Excel"
        assert payload["version"]["status"] == "draft"
        assert payload["version"]["version"] is None
        assert payload["version"]["cases"][0]["id"] == "imported-case"
        assert payload["version"]["cases"][0]["turns"][0]["id"] == "imported-turn"

        detail = client.get(f"/api/datasets/{dataset_id}")
        assert detail.status_code == 200
        assert detail.json()["dataset"]["description"] == "From workbook"

        published = client.post(f"/api/datasets/{dataset_id}/drafts/publish")
        assert published.status_code == 200
        downloaded = client.get(
            f"/api/datasets/{dataset_id}/versions/{published.json()['version']}/export/excel"
        )

        assert downloaded.status_code == 200
        assert downloaded.headers["content-type"] == XLSX_MEDIA_TYPE
        assert downloaded.headers["content-disposition"] == (
            'attachment; filename="Imported-Excel-v1.xlsx"'
        )
        sheet = openpyxl.load_workbook(BytesIO(downloaded.content), read_only=True)["Cases"]
        assert next(sheet.iter_rows(min_row=2, values_only=True))[0] == "imported-case"


def test_excel_export_rejects_a_draft_without_a_published_version(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={}),)),),
    ))

    with TestClient(create_app(tmp_path / "draft-excel-export-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Draft dataset"},
            files={"file": ("draft.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )
        assert imported.status_code == 201

        response = client.get(
            f"/api/datasets/{imported.json()['dataset']['id']}/versions/1/export/excel"
        )

        assert response.status_code == 404
        assert response.json()["detail"].startswith("unknown dataset version")


def test_excel_import_requires_a_name_and_creates_no_dataset(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={}),)),),
    ))

    with TestClient(create_app(tmp_path / "missing-name-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            files={"file": ("import.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _dataset_ids(client) == before


def test_excel_import_rejects_a_non_xlsx_filename_with_structured_issue(tmp_path):
    with TestClient(create_app(tmp_path / "wrong-extension-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Wrong extension"},
            files={"file": ("import.csv", b"not a spreadsheet", "text/csv")},
        )

        assert response.status_code == 422
        assert response.json()["detail"] == [{
            "sheet": "Cases",
            "row": None,
            "column": None,
            "message": "file must have a .xlsx filename",
        }]
        assert _dataset_ids(client) == before


def test_excel_import_rejects_bodies_over_ten_mebibytes_before_creating_a_dataset(tmp_path):
    with TestClient(create_app(tmp_path / "oversized-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Oversized"},
            files={
                "file": (
                    "oversized.xlsx",
                    b"x" * (10 * 1024 * 1024 + 1),
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert response.status_code == 422
        assert response.json()["detail"][0]["message"] == "XLSX upload exceeds 10 MiB"
        assert _dataset_ids(client) == before


def test_excel_import_returns_all_malformed_workbook_issues_and_creates_no_dataset(tmp_path):
    with TestClient(create_app(tmp_path / "malformed-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Malformed"},
            files={
                "file": (
                    "malformed.xlsx",
                    _workbook_with_multiple_issues(),
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert response.status_code == 422
        issues = response.json()["detail"]
        assert {issue["column"] for issue in issues} >= {
            "tags_json", "input_json", "expectations_json",
        }
        assert all(set(issue) == {"sheet", "row", "column", "message"} for issue in issues)
        assert _dataset_ids(client) == before


def test_excel_export_returns_not_found_for_an_unknown_published_version(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))

    with TestClient(create_app(tmp_path / "unknown-version-excel-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Published dataset"},
            files={"file": ("published.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )
        dataset_id = imported.json()["dataset"]["id"]
        assert client.post(f"/api/datasets/{dataset_id}/drafts/publish").status_code == 200

        response = client.get(f"/api/datasets/{dataset_id}/versions/2/export/excel")

        assert response.status_code == 404
        assert response.json()["detail"].startswith("unknown dataset version")
