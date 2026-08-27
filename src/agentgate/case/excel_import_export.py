"""XLSX import/export for Dataset cases."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
import json
from json import JSONDecodeError
import re
from typing import Any
from zipfile import BadZipFile, ZipFile
from xml.etree.ElementTree import ParseError

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from agentgate.domain import Case, DatasetVersion
from agentgate.domain.base import thaw_json

from .validation import ValidationIssue


SHEET_NAME = "Cases"
HEADERS = (
    "case_id", "case_name", "case_description", "category", "difficulty",
    "tags_json", "initial_state_json", "turn_id", "turn_order", "input_json",
    "expected_skill", "expectations_json", "required_tools_json", "forbidden_tools_json",
    "policy_rules_json", "turn_notes",
)
JSON_DEFAULTS: dict[str, Any] = {
    "tags_json": [],
    "initial_state_json": {},
    "input_json": {},
    "expectations_json": [],
    "required_tools_json": [],
    "forbidden_tools_json": [],
    "policy_rules_json": [],
}
OPTIONAL_DEFAULTS = {
    "case_description": "",
    "category": "positive",
    "difficulty": "medium",
    "expected_skill": None,
    "turn_notes": "",
}
REQUIRED_COLUMNS = ("case_name", "input_json")
REQUIRED_NON_JSON_COLUMNS = ("case_name",)
REQUIRED_HEADERS = ("case_name", "input_json")
MAX_INPUT_BYTES = 10 * 1024 * 1024
MAX_DATA_ROWS = 10_000
MAX_EXCEL_CELL_CHARACTERS = 32_767
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
_CONTENT_TYPES_PART = "[Content_Types].xml"
_PACKAGE_EXCEPTIONS = (
    BadZipFile,
    EOFError,
    InvalidFileException,
    KeyError,
    OSError,
    ParseError,
)
_DATASET_VALIDATION_PATH = re.compile(
    r"^cases\[(?P<case>\d+)\]"
    r"(?:\.turns\[(?P<turn>\d+)\])?"
    r"(?:\.(?P<field>.*))?$"
)


@dataclass(frozen=True, slots=True)
class ExcelImportIssue:
    sheet: str
    row: int | None
    column: str | None
    message: str


class DatasetExcelValidationError(ValueError):
    """All structural and model-validation errors found in an XLSX upload."""

    def __init__(self, issues: tuple[ExcelImportIssue, ...]) -> None:
        self.issues = issues
        details = "; ".join(
            f"{issue.sheet}!{issue.column or '*'}"
            f"{f' row {issue.row}' if issue.row is not None else ''}: {issue.message}"
            for issue in issues
        )
        super().__init__(details or "Excel workbook validation failed")


@dataclass(frozen=True, slots=True)
class ExcelExportIssue:
    sheet: str
    row: int | None
    column: str | None
    message: str


class DatasetExcelExportError(ValueError):
    """Cells that cannot be represented losslessly in an XLSX workbook."""

    def __init__(self, issues: tuple[ExcelExportIssue, ...]) -> None:
        self.issues = issues
        details = "; ".join(
            f"{issue.sheet}!{issue.column or '*'}"
            f"{f' row {issue.row}' if issue.row is not None else ''}: {issue.message}"
            for issue in issues
        )
        super().__init__(details or "Excel workbook export failed")


@dataclass(slots=True)
class _Row:
    number: int
    cells: dict[str, Any]
    turn_order: int | None


def excel_issues_from_dataset_validation(
    validation_issues: Iterable[ValidationIssue], cases: tuple[Case, ...],
) -> tuple[ExcelImportIssue, ...]:
    """Map Dataset validation paths back to their originating workbook cells."""
    row_starts: list[int] = []
    next_row = 2
    for case in cases:
        row_starts.append(next_row)
        next_row += len(case.turns)

    issues: list[ExcelImportIssue] = []
    for issue in validation_issues:
        if issue.path == "cases":
            issues.append(ExcelImportIssue(SHEET_NAME, None, None, issue.message))
            continue
        match = _DATASET_VALIDATION_PATH.fullmatch(issue.path)
        if match is None:
            issues.append(ExcelImportIssue(SHEET_NAME, None, None, issue.message))
            continue
        case_index = int(match.group("case"))
        if case_index >= len(cases):
            issues.append(ExcelImportIssue(SHEET_NAME, None, None, issue.message))
            continue
        turn_text = match.group("turn")
        turn_index = int(turn_text) if turn_text is not None else None
        if turn_index is not None and turn_index >= len(cases[case_index].turns):
            issues.append(ExcelImportIssue(SHEET_NAME, row_starts[case_index], None, issue.message))
            continue
        row = row_starts[case_index] + (turn_index or 0)
        field = match.group("field")
        column = _dataset_validation_column(field, turn_index is not None)
        issues.append(ExcelImportIssue(SHEET_NAME, row, column, issue.message))
    return tuple(issues)


def _dataset_validation_column(field: str | None, is_turn: bool) -> str | None:
    if not is_turn:
        return {
            "id": "case_id",
            "name": "case_name",
            "notes": "case_description",
            "tags": "tags_json",
            "initial_state": "initial_state_json",
        }.get(field or "")
    if field is None:
        return "required_tools_json"
    if field == "id":
        return "turn_id"
    if field == "input":
        return "input_json"
    if field.startswith("expectations["):
        return "expectations_json"
    return {
        "expected_skill": "expected_skill",
        "required_tools": "required_tools_json",
        "forbidden_tools": "forbidden_tools_json",
        "policy_rules": "policy_rules_json",
        "notes": "turn_notes",
    }.get(field)


def build_excel(version: DatasetVersion) -> bytes:
    """Serialize a DatasetVersion's Cases to the stable single-sheet XLSX format."""
    rows: list[tuple[Any, ...]] = []
    for case in version.cases:
        for turn_order, turn in enumerate(case.turns, start=1):
            rows.append((
                case.id,
                case.name,
                case.notes,
                case.category.value,
                case.difficulty.value,
                _compact_json(case.tags),
                _compact_json(case.initial_state),
                turn.id,
                turn_order,
                _compact_json(turn.input),
                turn.expected_skill,
                _compact_json([item.model_dump(mode="json") for item in turn.expectations]),
                _compact_json(turn.required_tools),
                _compact_json(turn.forbidden_tools),
                _compact_json(turn.policy_rules),
                turn.notes,
            ))
    export_issues = _excel_export_issues(rows)
    if export_issues:
        raise DatasetExcelExportError(tuple(export_issues))

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.append(_write_only_row(sheet, HEADERS))
    for row in rows:
        sheet.append(_write_only_row(sheet, row))
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _excel_export_issues(rows: list[tuple[Any, ...]]) -> list[ExcelExportIssue]:
    issues: list[ExcelExportIssue] = []
    for row_number, row in enumerate(rows, start=2):
        for column, value in zip(HEADERS, row, strict=True):
            if not isinstance(value, str):
                continue
            if len(value) > MAX_EXCEL_CELL_CHARACTERS:
                issues.append(ExcelExportIssue(
                    SHEET_NAME,
                    row_number,
                    column,
                    "cell text exceeds Excel's 32,767 character limit",
                ))
            if any(not _is_legal_xml_character(character) for character in value):
                issues.append(ExcelExportIssue(
                    SHEET_NAME,
                    row_number,
                    column,
                    "cell text contains a character that is not legal in XML",
                ))
    return issues


def _is_legal_xml_character(character: str) -> bool:
    codepoint = ord(character)
    return (
        codepoint in (0x09, 0x0A, 0x0D)
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def _write_only_row(sheet: Any, values: Iterable[Any]) -> tuple[Any, ...]:
    output: list[Any] = []
    for value in values:
        if isinstance(value, str):
            cell = WriteOnlyCell(sheet, value=value)
            cell.data_type = "s"
            output.append(cell)
        else:
            output.append(value)
    return tuple(output)


def parse_excel(content: bytes) -> tuple[Case, ...]:
    """Parse a Cases worksheet, collecting all independently discoverable issues."""
    if len(content) > MAX_INPUT_BYTES:
        raise DatasetExcelValidationError((ExcelImportIssue(
            SHEET_NAME, None, None, f"input exceeds {MAX_INPUT_BYTES} byte limit"
        ),))
    _validate_xlsx_archive(content)
    workbook = _load_xlsx_workbook(content)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise DatasetExcelValidationError((ExcelImportIssue(
                SHEET_NAME, None, None, f"required worksheet '{SHEET_NAME}' is missing"
            ),))
        sheet = workbook[SHEET_NAME]
        try:
            rows = iter(sheet.iter_rows())
        except _PACKAGE_EXCEPTIONS as exc:
            raise _invalid_xlsx_error(exc) from exc
        header_cells = _next_worksheet_row(rows) or ()
        header = tuple(cell.value for cell in header_cells)
        header_issues = _header_issues(header)
        header_issues.extend(
            ExcelImportIssue(
                SHEET_NAME,
                1,
                str(header[index]) if index < len(header) else None,
                "Excel formulas are not allowed",
            )
            for index, cell in enumerate(header_cells)
            if cell.data_type == "f"
        )
        if header_issues:
            raise DatasetExcelValidationError(tuple(header_issues))

        issues: list[ExcelImportIssue] = []
        cases: dict[str, list[_Row]] = {}
        anonymous_cases: list[list[_Row]] = []
        row_number = 2
        while (row_cells := _next_worksheet_row(rows)) is not None:
            if row_number > MAX_DATA_ROWS + 1:
                issues.append(ExcelImportIssue(
                    SHEET_NAME, row_number, None,
                    f"worksheet exceeds {MAX_DATA_ROWS} data row limit",
                ))
                break
            cells: dict[str, Any] = dict.fromkeys(HEADERS)
            for index, column in enumerate(header):
                cell = row_cells[index] if index < len(row_cells) else None
                if cell is None:
                    continue
                if cell.data_type == "f":
                    issues.append(ExcelImportIssue(
                        SHEET_NAME,
                        row_number,
                        column,
                        "Excel formulas are not allowed",
                    ))
                    cells[str(column)] = _formula_fallback(str(column), row_number)
                else:
                    cells[str(column)] = cell.value
            if all(_is_blank(value) for value in cells.values()):
                row_number += 1
                continue
            parsed = _parse_row(cells, row_number, issues)
            case_id = parsed.cells["case_id"]
            if not _is_blank(case_id):
                cases.setdefault(str(case_id), []).append(parsed)
            else:
                anonymous_cases.append([parsed])
            row_number += 1

        output: list[Case] = []
        for case_rows in [*cases.values(), *anonymous_cases]:
            _validate_case_structure(case_rows, issues)
            parsed_case = _build_case(case_rows, issues)
            if parsed_case is not None:
                output.append(parsed_case)

        if issues:
            raise DatasetExcelValidationError(tuple(issues))
        return tuple(output)
    finally:
        workbook.close()


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            names = {member.filename for member in members}
    except (BadZipFile, OSError) as exc:
        raise _invalid_xlsx_error(exc) from exc

    issues: list[ExcelImportIssue] = []
    total_uncompressed = sum(member.file_size for member in members)
    for member in members:
        if member.file_size > MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES:
            issues.append(ExcelImportIssue(
                SHEET_NAME,
                None,
                None,
                f"archive entry '{member.filename}' exceeds the "
                f"{MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES // (1024 * 1024)} MiB "
                "single-entry uncompressed limit",
            ))
        ratio = member.file_size / max(member.compress_size, 1)
        if ratio > MAX_XLSX_COMPRESSION_RATIO:
            issues.append(ExcelImportIssue(
                SHEET_NAME,
                None,
                None,
                f"archive entry '{member.filename}' exceeds the "
                f"{MAX_XLSX_COMPRESSION_RATIO}:1 compression ratio limit",
            ))
    if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        issues.append(ExcelImportIssue(
            SHEET_NAME,
            None,
            None,
            f"archive exceeds the {MAX_XLSX_UNCOMPRESSED_BYTES // (1024 * 1024)} MiB "
            "total uncompressed limit",
        ))
    if issues:
        raise DatasetExcelValidationError(tuple(issues))
    if _CONTENT_TYPES_PART not in names:
        raise DatasetExcelValidationError((ExcelImportIssue(
            SHEET_NAME,
            None,
            None,
            f"invalid XLSX content: missing {_CONTENT_TYPES_PART}",
        ),))


def _load_xlsx_workbook(content: bytes) -> Any:
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=False)
    except _PACKAGE_EXCEPTIONS as exc:
        raise _invalid_xlsx_error(exc) from exc


def _next_worksheet_row(rows: Any) -> tuple[Any, ...] | None:
    try:
        return tuple(next(rows))
    except StopIteration:
        return None
    except _PACKAGE_EXCEPTIONS as exc:
        raise _invalid_xlsx_error(exc) from exc


def _invalid_xlsx_error(exc: BaseException) -> DatasetExcelValidationError:
    return DatasetExcelValidationError((ExcelImportIssue(
        SHEET_NAME, None, None, f"invalid XLSX content: {exc}"
    ),))


def _formula_fallback(column: str, row: int) -> Any:
    return {
        "case_id": f"invalid-formula-case-{row}",
        "case_name": "Invalid formula",
        "case_description": "",
        "category": "positive",
        "difficulty": "medium",
        "tags_json": "[]",
        "initial_state_json": "{}",
        "turn_id": f"invalid-formula-turn-{row}",
        "turn_order": 1,
        "input_json": "{}",
        "expected_skill": None,
        "expectations_json": "[]",
        "required_tools_json": "[]",
        "forbidden_tools_json": "[]",
        "policy_rules_json": "[]",
        "turn_notes": "",
    }[column]


def _compact_json(value: Any) -> str:
    return json.dumps(thaw_json(value), ensure_ascii=False, separators=(",", ":"))


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _header_issues(header: tuple[Any, ...]) -> list[ExcelImportIssue]:
    issues: list[ExcelImportIssue] = []
    seen: set[str] = set()
    for actual in header:
        if actual not in HEADERS:
            issues.append(ExcelImportIssue(
                SHEET_NAME, 1, str(actual) if actual is not None else None,
                "unexpected header",
            ))
            continue
        if actual in seen:
            issues.append(ExcelImportIssue(
                SHEET_NAME, 1, str(actual), "duplicate header",
            ))
        seen.add(str(actual))
    for required in REQUIRED_HEADERS:
        if required not in seen:
            issues.append(ExcelImportIssue(
                SHEET_NAME, 1, required, f"required header '{required}' is missing",
            ))
    return issues


def _parse_row(
    raw_cells: dict[str, Any], row: int, issues: list[ExcelImportIssue],
) -> _Row:
    cells = dict(raw_cells)
    for column, default in OPTIONAL_DEFAULTS.items():
        if _is_blank(cells[column]):
            cells[column] = default
    for column in REQUIRED_NON_JSON_COLUMNS:
        if _is_blank(cells[column]):
            issues.append(ExcelImportIssue(SHEET_NAME, row, column, "value is required"))

    for column, default in JSON_DEFAULTS.items():
        required = column in REQUIRED_COLUMNS
        cells[column] = _parse_json_cell(cells[column], column, row, required, default, issues)
    turn_order = _parse_turn_order(cells["turn_order"], row, issues)
    return _Row(row, cells, turn_order)


def _parse_json_cell(
    value: Any,
    column: str,
    row: int,
    required: bool,
    default: Any,
    issues: list[ExcelImportIssue],
) -> Any:
    if _is_blank(value):
        if required:
            issues.append(ExcelImportIssue(SHEET_NAME, row, column, "value is required"))
        return default
    if not isinstance(value, str):
        issues.append(ExcelImportIssue(SHEET_NAME, row, column, "must contain JSON text"))
        return default
    try:
        return json.loads(value)
    except JSONDecodeError as exc:
        issues.append(ExcelImportIssue(SHEET_NAME, row, column, f"invalid JSON: {exc.msg}"))
        return default


def _parse_turn_order(value: Any, row: int, issues: list[ExcelImportIssue]) -> int | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        issues.append(ExcelImportIssue(SHEET_NAME, row, "turn_order", "must be an integer"))
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            pass
    issues.append(ExcelImportIssue(SHEET_NAME, row, "turn_order", "must be an integer"))
    return None


def _validate_case_structure(rows: list[_Row], issues: list[ExcelImportIssue]) -> None:
    first = rows[0]
    for row in rows[1:]:
        for column in (
            "case_name", "case_description", "category", "difficulty", "tags_json",
            "initial_state_json",
        ):
            if row.cells[column] != first.cells[column]:
                issues.append(ExcelImportIssue(
                    SHEET_NAME, row.number, column,
                    f"conflicts with the first row for case_id '{first.cells['case_id']}'",
                ))

    seen_turn_ids: set[str] = set()
    seen_orders: set[int] = set()
    provided_orders = [row for row in rows if row.turn_order is not None]
    if not provided_orders:
        for order, row in enumerate(rows, start=1):
            row.turn_order = order
    elif len(provided_orders) != len(rows):
        for row in rows:
            if row.turn_order is None:
                issues.append(ExcelImportIssue(
                    SHEET_NAME, row.number, "turn_order",
                    "value is required when another turn order is provided",
                ))
    for row in rows:
        turn_id = row.cells["turn_id"]
        if not _is_blank(turn_id):
            if str(turn_id) in seen_turn_ids:
                issues.append(ExcelImportIssue(
                    SHEET_NAME, row.number, "turn_id", "duplicate turn_id within Case"
                ))
            seen_turn_ids.add(str(turn_id))
        if row.turn_order is not None:
            if row.turn_order in seen_orders:
                issues.append(ExcelImportIssue(
                    SHEET_NAME, row.number, "turn_order", "duplicate turn_order within Case"
                ))
            seen_orders.add(row.turn_order)

    ordered = sorted((row.turn_order for row in rows if row.turn_order is not None))
    if len(ordered) == len(rows) and ordered != list(range(1, len(rows) + 1)):
        valid_orders = range(1, len(rows) + 1)
        offender = next((row for row in rows if row.turn_order not in valid_orders), rows[-1])
        issues.append(ExcelImportIssue(
            SHEET_NAME, offender.number, "turn_order", "must be 1-based and contiguous per Case"
        ))


def _build_case(rows: list[_Row], issues: list[ExcelImportIssue]) -> Case | None:
    ordered_rows = sorted(rows, key=lambda row: (
        row.turn_order is None,
        row.turn_order if row.turn_order is not None else row.number,
    ))
    first = ordered_rows[0]
    payload: dict[str, Any] = {
        "name": first.cells["case_name"],
        "notes": first.cells["case_description"],
        "category": first.cells["category"],
        "difficulty": first.cells["difficulty"],
        "tags": first.cells["tags_json"],
        "initial_state": first.cells["initial_state_json"],
        "turns": [
            {
                "input": row.cells["input_json"],
                "expected_skill": row.cells["expected_skill"],
                "expectations": row.cells["expectations_json"],
                "required_tools": row.cells["required_tools_json"],
                "forbidden_tools": row.cells["forbidden_tools_json"],
                "policy_rules": row.cells["policy_rules_json"],
                "notes": row.cells["turn_notes"],
            }
            for row in ordered_rows
        ],
    }
    if not _is_blank(first.cells["case_id"]):
        payload["id"] = first.cells["case_id"]
    for turn_payload, row in zip(payload["turns"], ordered_rows, strict=True):
        if not _is_blank(row.cells["turn_id"]):
            turn_payload["id"] = row.cells["turn_id"]
    try:
        return Case.model_validate(payload)
    except ValidationError as exc:
        for error in exc.errors():
            row, column = _validation_location(error["loc"], first.number, ordered_rows)
            issues.append(ExcelImportIssue(SHEET_NAME, row, column, error["msg"]))
        return None


def _validation_location(
    location: Iterable[Any], first_row: int, rows: list[_Row],
) -> tuple[int, str | None]:
    parts = tuple(location)
    case_columns = {
        "id": "case_id", "name": "case_name", "notes": "case_description",
        "category": "category", "difficulty": "difficulty", "tags": "tags_json",
        "initial_state": "initial_state_json",
    }
    turn_columns = {
        "id": "turn_id", "input": "input_json", "expected_skill": "expected_skill",
        "expectations": "expectations_json", "required_tools": "required_tools_json",
        "forbidden_tools": "forbidden_tools_json", "policy_rules": "policy_rules_json",
        "notes": "turn_notes",
    }
    if len(parts) >= 3 and parts[0] == "turns" and isinstance(parts[1], int):
        return rows[parts[1]].number, turn_columns.get(str(parts[2]))
    return first_row, case_columns.get(str(parts[0])) if parts else None
